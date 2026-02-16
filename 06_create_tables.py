#!/usr/bin/env python3
from __future__ import annotations

import argparse
import hashlib
import json
import re
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Optional

from datetime import date
from zoneinfo import ZoneInfo
import datetime as _dt

import requests
from tqdm import tqdm
from urllib.parse import urljoin
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from crz_config import OUT_JSON_DIR, ATT_DIR
from helper import squash_spaced_letters

from crz_page import fetch_and_parse_crz_id, fetch_and_parse_crz_url, ascii_key


# ----------------------------
# misc helpers
# ----------------------------

_RX_ISO = re.compile(r"^\s*(\d{4})-(\d{2})-(\d{2})(?:\s+.*)?$")
_RX_DMY = re.compile(r"^\s*(\d{1,2})\.(\d{1,2})\.(\d{4})\s*$")

def today_bratislava() -> date:
    return _dt.datetime.now(ZoneInfo("Europe/Bratislava")).date()

def parse_date_any(v: Any) -> Optional[date]:
    """
    Returns a date or None.
    None means "neuvedený" / "0000-00-00" / empty / unparseable => treat as unlimited validity.
    Accepts:
      - YYYY-MM-DD
      - YYYY-MM-DD HH:MM:SS (or anything after the date)
      - D.M.YYYY / DD.MM.YYYY
    """
    if v is None or isinstance(v, (int, float)):
        return None

    s = str(v).strip()
    if not s or s == "0000-00-00":
        return None

    m = _RX_ISO.match(s)
    if m:
        try:
            y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
            if y == 0 or mo == 0 or d == 0:
                return None
            return date(y, mo, d)
        except Exception:
            return None

    m = _RX_DMY.match(s)
    if m:
        try:
            d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
            return date(y, mo, d)
        except Exception:
            return None

    return None

def is_expired(v: Any, *, today: date) -> bool:
    d = parse_date_any(v)
    # "today is ok, just not before"
    return (d is not None) and (d < today)

def fmt_date(v: Any) -> str:
    d = parse_date_any(v)
    if d is None:
        return "neuvedený"
    return f"{d.day:02d}.{d.month:02d}.{d.year:04d}"


def digits_only(s: str) -> str:
    return re.sub(r"\D+", "", s or "")


def strip_ico(s: Any) -> str:
    """
    Normalize ICO for comparisons:
    - keep digits only
    - strip leading zeros
    """
    d = digits_only(str(s or ""))
    d2 = d.lstrip("0")
    return d2 if d2 else d  # keep "0" if all zeros


def _norm_url(page_url: str, href: str) -> str:
    href = (href or "").strip()
    if not href:
        return ""
    return urljoin(page_url or "https://crz.gov.sk/", href)


def _att_blob(a: dict[str, Any]) -> str:
    parts = [
        a.get("text"),
        a.get("label"),
        a.get("title"),
        a.get("aria_label"),
        a.get("filename"),
        a.get("file_kind"),
        a.get("size_text"),
    ]
    return " ".join(str(p or "") for p in parts).lower()


LINK_FONT = Font(color="0563C1", underline="single")

def apply_link(cell, url: str) -> None:
    cell.hyperlink = url
    cell.font = LINK_FONT


def pick_attachment(
    page: Dict[str, Any],
    *,
    want: str,  # "text" or "scan"
) -> tuple[str, Optional[str]]:
    """
    Returns (cell_text, hyperlink_url or None).

    - want="text": choose attachment whose label/text/aria contains "textov"
    - want="scan": prefer "podpis" / "podpís" / "scan"/"sken", else any non-text attachment
    """
    prilohy = page.get("prilohy") or []
    if not isinstance(prilohy, list):
        return ("", None)

    page_url = str(page.get("url") or "https://crz.gov.sk/")

    cand: list[dict[str, Any]] = []
    for a in prilohy:
        if not isinstance(a, dict):
            continue
        href = a.get("href") or a.get("url")
        if isinstance(href, str) and href.strip():
            cand.append(a)

    if not cand:
        return ("", None)

    def cell_text_for(a: dict[str, Any]) -> str:
        txt = str(a.get("text") or a.get("label") or a.get("title") or "").strip()
        if txt:
            return txt
        fn = str(a.get("filename") or "").strip()
        if fn:
            return fn
        href = str(a.get("href") or a.get("url") or "")
        return href.rsplit("/", 1)[-1]

    text_cand = [a for a in cand if "textov" in _att_blob(a)]
    nontext_cand = [a for a in cand if a not in text_cand]

    if want == "text":
        if not text_cand:
            return ("", None)
        a = text_cand[0]
        href = _norm_url(page_url, str(a.get("href") or a.get("url") or ""))
        return (cell_text_for(a), href or None)

    if want == "scan":
        if nontext_cand:
            def score(a: dict[str, Any]) -> int:
                b = _att_blob(a)
                s = 0
                if "podpis" in b or "podpís" in b or "podpisan" in b:
                    s += 10
                if "sken" in b or "scan" in b:
                    s += 5
                return s
            nontext_cand.sort(key=score, reverse=True)
            a = nontext_cand[0]
        else:
            a = text_cand[0]

        href = _norm_url(page_url, str(a.get("href") or a.get("url") or ""))
        return (cell_text_for(a), href or None)

    return ("", None)


def classify_page(page: Dict[str, Any]) -> str:
    """
    contract vs amendment detection
    """
    ident = page.get("identifikacia") or {}
    ident_ascii = page.get("identifikacia_ascii") or {}

    if isinstance(ident_ascii, dict) and any(k in ident_ascii for k in ("cislo_dodatku", "id_dodatku")):
        return "amendment"
    if isinstance(ident, dict) and any("dodatku" in str(k).lower() for k in ident.keys()):
        return "amendment"

    typ = ""
    if isinstance(ident_ascii, dict):
        typ = ident_ascii.get("typ") or ""
    if not typ and isinstance(ident, dict):
        typ = ident.get("Typ") or ""

    t = str(typ).strip().lower()
    if t:
        return "amendment" if "dodatok" in t else "contract"

    title = str(page.get("title") or "").lower()
    return "amendment" if title.startswith("dodatok") else "contract"


def datum_ascii_map(page: Dict[str, Any]) -> dict[str, str]:
    d = page.get("datum") or {}
    if not isinstance(d, dict):
        return {}
    out: dict[str, str] = {}
    for k, v in d.items():
        if not isinstance(k, str):
            continue
        out[ascii_key(k)] = str(v or "")
    return out


def get_page_field(page: Dict[str, Any], *, ident_key: str) -> str:
    ident_ascii = page.get("identifikacia_ascii") or {}
    if isinstance(ident_ascii, dict):
        v = ident_ascii.get(ident_key)
        if v is not None:
            return str(v)
    return ""


def parse_addendum_numeric_id(page: Dict[str, Any], addendum_url: str) -> Optional[int]:
    pid = page.get("page_id")
    if isinstance(pid, int):
        return pid
    m = re.search(r"/(\d+)(?:[/-]|$)", addendum_url)
    if m:
        try:
            return int(m.group(1))
        except Exception:
            return None
    return None


# ----------------------------
# reading filtered.jsonl
# ----------------------------

def load_filtered_lines(filtered_jsonl: Path) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    with open(filtered_jsonl, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                out.append(json.loads(line))
            except Exception:
                continue
    return out


def parse_ico_allow(args) -> set[str]:
    allowed: set[str] = set()

    if args.ico_allow:
        parts = [p.strip() for p in args.ico_allow.split(",") if p.strip()]
        allowed.update(strip_ico(p) for p in parts if strip_ico(p))

    if args.ico_allow_file:
        p = Path(args.ico_allow_file)
        if p.exists():
            for line in p.read_text(encoding="utf-8").splitlines():
                s = strip_ico(line.strip())
                if s:
                    allowed.add(s)

    return allowed


# ----------------------------
# restartable state: rows.jsonl + done_contracts.jsonl
# ----------------------------

def append_jsonl(path: Path, obj: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "a", encoding="utf-8") as f:
        f.write(json.dumps(obj, ensure_ascii=False) + "\n")


def sha1_short(s: str) -> str:
    return hashlib.sha1(s.encode("utf-8")).hexdigest()[:16]


def load_state(rows_jsonl: Path, done_jsonl: Path) -> tuple[set[int], set[str]]:
    done_cids: set[int] = set()
    seen_row_keys: set[str] = set()

    if done_jsonl.exists():
        for line in done_jsonl.read_text(encoding="utf-8").splitlines():
            if not line.strip():
                continue
            try:
                o = json.loads(line)
            except Exception:
                continue
            if o.get("kind") == "done_contract":
                try:
                    done_cids.add(int(o.get("cid")))
                except Exception:
                    pass

    if rows_jsonl.exists():
        for line in rows_jsonl.read_text(encoding="utf-8").splitlines():
            if not line.strip():
                continue
            try:
                o = json.loads(line)
            except Exception:
                continue
            if o.get("kind") == "row":
                rk = o.get("row_key")
                if isinstance(rk, str):
                    seen_row_keys.add(rk)

    return done_cids, seen_row_keys


def write_row_state(
    rows_jsonl: Path,
    *,
    seen_row_keys: set[str],
    row_key: str,
    contract_ord: int,
    sub: int,
    parent_cid: int,
    values: list[Any],
    links: dict[str, Optional[str]],
    comments: dict[str, str],
) -> bool:
    if row_key in seen_row_keys:
        return False
    append_jsonl(rows_jsonl, {
        "kind": "row",
        "ts": time.time(),
        "row_key": row_key,
        "contract_ord": contract_ord,
        "sub": sub,
        "parent_cid": parent_cid,
        "values": values,
        "links": links,
        "comments": comments,
    })
    seen_row_keys.add(row_key)
    return True


def mark_done(done_jsonl: Path, *, done_cids: set[int], cid: int) -> None:
    if cid in done_cids:
        return
    append_jsonl(done_jsonl, {"kind": "done_contract", "ts": time.time(), "cid": cid})
    done_cids.add(cid)


# ----------------------------
# Excel rebuild from rows.jsonl
# ----------------------------

HEADERS = [
    "Dodávateľ",
    "IČO dodávateľa",
    "Názov zmluvy",
    "Číslo zmluvy",
    "Dátum zverejnenia",
    "Dátum uzavretia",
    "Dátum účinnosti",
    "Dátum platnosti do",
    "Objednávateľ",
    "IČO objednávateľa",
    "ID zmluvy",
    "Sken prílohy",
    "Text prílohy",
]

FILL_A = PatternFill("solid", fgColor="E8F5E9")   # light green
FILL_B = PatternFill("solid", fgColor="FFFDE7")   # light yellow
HEADER_FILL = PatternFill("solid", fgColor="ECEFF1")  # light gray-blue

def apply_row_fill(ws, row: int, max_col: int, fill: PatternFill) -> None:
    for c in range(1, max_col + 1):
        ws.cell(row=row, column=c).fill = fill

def set_comment(ws, row: int, col: int, text: str) -> None:
    cell = ws.cell(row=row, column=col)
    cell.comment = Comment(text, "crz_scraper")


def build_xlsx_from_rows(rows_jsonl: Path, out_xlsx: Path) -> None:
    rows: list[dict[str, Any]] = []
    if rows_jsonl.exists():
        for line in rows_jsonl.read_text(encoding="utf-8").splitlines():
            if not line.strip():
                continue
            try:
                o = json.loads(line)
            except Exception:
                continue
            if o.get("kind") == "row":
                rows.append(o)

    rows.sort(key=lambda r: (int(r.get("contract_ord", 10**12)), int(r.get("sub", 10**12))))

    wb = Workbook()
    ws = wb.active
    ws.title = "contracts"

    # header
    for col, h in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"

    wrap_cols = {3, 12, 13}
    for col in range(1, len(HEADERS) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 22
        if col in (3, 12, 13):
            ws.column_dimensions[get_column_letter(col)].width = 45
        if col == 11:
            ws.column_dimensions[get_column_letter(col)].width = 12

    row_idx = 2
    group_fill = FILL_A
    cur_ord: int | None = None

    for r in rows:
        ord_ = int(r.get("contract_ord", -1))
        if cur_ord is None:
            cur_ord = ord_
        elif ord_ != cur_ord:
            group_fill = FILL_B if group_fill == FILL_A else FILL_A
            cur_ord = ord_

        values = r.get("values") or []
        links = r.get("links") or {}
        comments = r.get("comments") or {}

        for col, v in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col, value=v)
            cell.alignment = Alignment(vertical="top", wrap_text=(col in wrap_cols))
        apply_row_fill(ws, row_idx, len(HEADERS), group_fill)

        # hyperlinks: id col=11, scan col=12, text col=13
        if isinstance(links, dict):
            if links.get("id"):
                apply_link(ws.cell(row=row_idx, column=11), str(links["id"]))
            if links.get("scan"):
                apply_link(ws.cell(row=row_idx, column=12), str(links["scan"]))
            if links.get("text"):
                apply_link(ws.cell(row=row_idx, column=13), str(links["text"]))

        # comments: B=2 supplier ico, J=10 buyer ico
        if isinstance(comments, dict):
            if comments.get("B"):
                set_comment(ws, row_idx, 2, str(comments["B"]))
            if comments.get("J"):
                set_comment(ws, row_idx, 10, str(comments["J"]))

        row_idx += 1

    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{max(1, row_idx - 1)}"
    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_xlsx)


# ----------------------------
# dump reader cache (OUT_JSON_DIR/<file>.json)
# ----------------------------

def get_recs_for_file_cached_factory():
    last_file: Optional[str] = None
    last_recs: Optional[list[dict[str, Any]]] = None

    def get_recs_for_file(fn: str) -> Optional[list[dict[str, Any]]]:
        nonlocal last_file, last_recs
        if fn == last_file and last_recs is not None:
            return last_recs

        p = OUT_JSON_DIR / fn
        try:
            obj = json.loads(p.read_text(encoding="utf-8"))
        except Exception:
            last_file, last_recs = fn, None
            return None

        z = obj.get("zmluva") if isinstance(obj, dict) else None
        if isinstance(z, dict):
            recs = [z]
        elif isinstance(z, list):
            recs = [r for r in z if isinstance(r, dict)]
        else:
            recs = None

        last_file, last_recs = fn, recs
        return recs

    return get_recs_for_file


# ----------------------------
# main
# ----------------------------

def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--run-dir", type=str, default="", help="filtered_contracts/<sha>/ directory")
    ap.add_argument("--filtered-jsonl", type=str, default="", help="Path to filtered.jsonl (overrides --run-dir)")
    ap.add_argument("--out", type=str, default="", help="Output xlsx path (default: <run-dir>/table.xlsx)")

    ap.add_argument("--ico-allow", type=str, default="", help="Comma-separated supplier ICO allowlist (leading zeros ignored)")
    ap.add_argument("--ico-allow-file", type=str, default="", help="Text file with allowed ICOs (one per line)")

    ap.add_argument("--min-interval-s", type=float, default=0.0, help="Throttle between network requests (CRZ fetch)")
    ap.add_argument("--refetch-days", type=float, default=7.0, help="Re-fetch cached pages older than N days")

    ap.add_argument("--resume", action="store_true", help="Resume using rows.jsonl + done_contracts.jsonl")
    ap.add_argument("--write-xlsx-every", type=int, default=0,
                    help="Rebuild XLSX every N finished contracts (0=only at end)")

    ap.add_argument("--verbose", action="store_true")
    args = ap.parse_args()

    TODAY = today_bratislava()

    # resolve input paths
    if args.filtered_jsonl:
        filtered_jsonl = Path(args.filtered_jsonl)
        run_root = filtered_jsonl.parent
    else:
        if not args.run_dir:
            raise SystemExit("Provide --run-dir filtered_contracts/<sha>/ or --filtered-jsonl ...")
        run_root = Path(args.run_dir)
        filtered_jsonl = run_root / "filtered.jsonl"

    if not filtered_jsonl.exists():
        raise SystemExit(f"Missing {filtered_jsonl}")

    out_xlsx = Path(args.out) if args.out else (run_root / "table.xlsx")
    rows_jsonl = run_root / "rows.jsonl"
    done_jsonl = run_root / "done_contracts.jsonl"

    allowed_icos = parse_ico_allow(args)

    done_cids, seen_row_keys = load_state(rows_jsonl, done_jsonl) if args.resume else (set(), set())
    sess = requests.Session()

    lines = load_filtered_lines(filtered_jsonl)
    get_recs_for_file = get_recs_for_file_cached_factory()

    # stats for progress
    n_skipped_resume = 0
    n_skipped_allow_early = 0
    n_notfound = 0
    n_amendment = 0
    n_written_rows = len(seen_row_keys)
    n_done = len(done_cids)

    pbar = tqdm(enumerate(lines), total=len(lines), desc="Contracts", unit="rec")

    for ord_i, obj in pbar:
        ref = obj.get("ref") or {}
        fn = ref.get("file")
        idx = ref.get("index")
        if not isinstance(fn, str):
            continue
        try:
            idx_i = int(idx)
        except Exception:
            continue

        recs = get_recs_for_file(fn)
        if not recs or idx_i < 0 or idx_i >= len(recs):
            continue
        rec = recs[idx_i]

        if is_expired(rec.get("datum_platnost_do"), today=TODAY):
            continue

        # base contract id
        try:
            cid = int(rec.get("ID"))
        except Exception:
            try:
                cid = int(obj.get("id"))
            except Exception:
                continue

        # resume skip
        if args.resume and cid in done_cids:
            n_skipped_resume += 1
            if (n_skipped_resume % 50) == 0:
                pbar.set_postfix_str(
                    f"done={n_done} rows={n_written_rows} skip_resume={n_skipped_resume} nf={n_notfound}"
                )
            continue

        # allowlist EARLY check if ICO present in dump (saves network)
        supplier_ico_dump = strip_ico(rec.get("ico"))
        if allowed_icos and supplier_ico_dump and supplier_ico_dump not in allowed_icos:
            n_skipped_allow_early += 1
            continue

        # fetch+parse page
        page = fetch_and_parse_crz_id(
            cid,
            sleep_s=0.0,
            session=sess,
            min_interval_s=float(args.min_interval_s),
            refetch_days=float(args.refetch_days),
        )

        if not page.get("found"):
            n_notfound += 1
            if args.verbose:
                print(f"[skip] id={cid}: CRZ page not found: {page.get('warnings')}")
            continue

        if classify_page(page) == "amendment":
            n_amendment += 1
            if args.verbose:
                print(f"[skip] base id={cid} is an amendment according to CRZ page")
            continue

        # supplier ICO (dump preferred, page fallback)
        supplier_ico = supplier_ico_dump or strip_ico(get_page_field(page, ident_key="ico_dodavatela"))

        # allowlist LATE check if dump ICO was missing
        if allowed_icos and supplier_ico and supplier_ico not in allowed_icos:
            continue

        # contract row data (prefer dump, fallback to page)
        supplier_name = squash_spaced_letters(str(rec.get("zs2") or "")) or get_page_field(page, ident_key="dodavatel")
        buyer_name = squash_spaced_letters(str(rec.get("zs1") or "")) or get_page_field(page, ident_key="objednavatel")

        buyer_ico = strip_ico(rec.get("ico1")) or strip_ico(get_page_field(page, ident_key="ico_objednavatela"))

        predmet = squash_spaced_letters(str(rec.get("predmet") or "")) or str(page.get("title") or "")
        cislo = str(rec.get("nazov") or "") or get_page_field(page, ident_key="cislo_zmluvy")

        d_zver = fmt_date(rec.get("datum_zverejnene"))
        d_uzav = fmt_date(rec.get("potv_datum"))
        d_ucin = fmt_date(rec.get("datum_ucinnost"))
        d_plat = fmt_date(rec.get("datum_platnost_do"))

        scan_text, scan_url = pick_attachment(page, want="scan")
        text_text, text_url = pick_attachment(page, want="text")

        # write contract row to STATE
        values = [
            supplier_name,
            supplier_ico,
            predmet,
            cislo,
            d_zver,
            d_uzav,
            d_ucin,
            d_plat,
            buyer_name,
            buyer_ico,
            cid,
            scan_text,
            text_text,
        ]

        wrote = write_row_state(
            rows_jsonl,
            seen_row_keys=seen_row_keys,
            row_key=f"c:{cid}",
            contract_ord=ord_i,
            sub=0,
            parent_cid=cid,
            values=values,
            links={"id": str(page.get("url") or ""), "scan": scan_url, "text": text_url},
            comments={},
        )
        if wrote:
            n_written_rows += 1

        # addenda
        dodatky = page.get("dodatky") or []
        if isinstance(dodatky, list) and dodatky:
            base = str(page.get("url") or "https://crz.gov.sk/")
            for add_i, dd in enumerate(dodatky):
                if not isinstance(dd, dict):
                    continue
                add_url = dd.get("addendum_url")
                if not isinstance(add_url, str) or not add_url:
                    continue

                add_url_abs = urljoin(base, add_url)

                add_page = fetch_and_parse_crz_url(
                    add_url_abs,
                    session=sess,
                    min_interval_s=float(args.min_interval_s),
                    refetch_days=float(args.refetch_days),
                )
                if not add_page.get("found"):
                    if args.verbose:
                        print(f"[skip] addendum url={add_url_abs}: not found: {add_page.get('warnings')}")
                    continue

                add_id_num = parse_addendum_numeric_id(add_page, add_url_abs)
                if add_id_num is not None:
                    add_row_key = f"a:{add_id_num}"
                    add_id_cell = add_id_num
                else:
                    add_row_key = f"au:{sha1_short(add_url_abs)}"
                    add_id_cell = ""

                add_ident_supplier = get_page_field(add_page, ident_key="dodavatel") or supplier_name
                add_ident_buyer = get_page_field(add_page, ident_key="objednavatel") or buyer_name
                add_ico_supplier = strip_ico(get_page_field(add_page, ident_key="ico_dodavatela")) or supplier_ico
                add_ico_buyer = strip_ico(get_page_field(add_page, ident_key="ico_objednavatela")) or buyer_ico

                dmap = datum_ascii_map(add_page)
                add_d_zver = fmt_date(dmap.get("datum_zverejnenia"))
                add_d_uzav = fmt_date(dmap.get("datum_uzavretia"))
                add_d_ucin = fmt_date(dmap.get("datum_ucinnosti"))
                add_d_plat = fmt_date(dmap.get("datum_platnosti_do"))

                add_title = str(add_page.get("title") or dd.get("addendum_title") or "").strip()
                add_num = (
                    get_page_field(add_page, ident_key="cislo_dodatku")
                    or get_page_field(add_page, ident_key="cislo_dodatku_2")
                    or str(dd.get("addendum_number") or "")
                )

                add_scan_text, add_scan_url = pick_attachment(add_page, want="scan")
                add_text_text, add_text_url = pick_attachment(add_page, want="text")

                values2 = [
                    add_ident_supplier,
                    add_ico_supplier,
                    add_title,
                    add_num,
                    add_d_zver,
                    add_d_uzav,
                    add_d_ucin,
                    add_d_plat,
                    add_ident_buyer,
                    add_ico_buyer,
                    add_id_cell,
                    add_scan_text,
                    add_text_text,
                ]

                comments: dict[str, str] = {}
                if supplier_ico and add_ico_supplier and supplier_ico != add_ico_supplier:
                    comments["B"] = f"Supplier ICO mismatch: parent={supplier_ico} addendum={add_ico_supplier}"
                if buyer_ico and add_ico_buyer and buyer_ico != add_ico_buyer:
                    comments["J"] = f"Buyer ICO mismatch: parent={buyer_ico} addendum={add_ico_buyer}"

                wrote2 = write_row_state(
                    rows_jsonl,
                    seen_row_keys=seen_row_keys,
                    row_key=add_row_key,
                    contract_ord=ord_i,
                    sub=1 + add_i,
                    parent_cid=cid,
                    values=values2,
                    links={"id": str(add_page.get("url") or add_url_abs), "scan": add_scan_url, "text": add_text_url},
                    comments=comments,
                )
                if wrote2:
                    n_written_rows += 1

        # mark contract done (makes resume work)
        mark_done(done_jsonl, done_cids=done_cids, cid=cid)
        n_done += 1

        # optional periodic xlsx rebuild
        if args.write_xlsx_every and (n_done % int(args.write_xlsx_every) == 0):
            build_xlsx_from_rows(rows_jsonl, out_xlsx)

        # progress info
        pbar.set_postfix_str(
            f"done={n_done} rows={n_written_rows} skip_resume={n_skipped_resume} "
            f"skip_allow={n_skipped_allow_early} nf={n_notfound} amend={n_amendment}"
        )

    # final xlsx rebuild
    build_xlsx_from_rows(rows_jsonl, out_xlsx)
    print(f"[ok] wrote {out_xlsx}")
    print(f"[state] rows={rows_jsonl} done={done_jsonl}")


if __name__ == "__main__":
    main()
