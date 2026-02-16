#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import re
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Optional

from datetime import date
from zoneinfo import ZoneInfo
import datetime as _dt

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from helper import find_project_root

# ----------------------------
# shared-ish helpers
# ----------------------------

_RX_ISO = re.compile(r"^\s*(\d{4})-(\d{2})-(\d{2})(?:\s+.*)?$")
_RX_DMY = re.compile(r"^\s*(\d{1,2})\.(\d{1,2})\.(\d{4})\s*$")


def today_bratislava() -> date:
    return _dt.datetime.now(ZoneInfo("Europe/Bratislava")).date()


def parse_date_any(v: Any) -> Optional[date]:
    """
    Returns a date or None.
    None means unlimited/unknown: "neuvedený" / "0000-00-00" / empty / unparseable
    Accepts:
      - YYYY-MM-DD
      - YYYY-MM-DD HH:MM:SS (or anything after the date)
      - D.M.YYYY / DD.MM.YYYY
    """
    if v is None or isinstance(v, (int, float)):
        return None

    s = str(v).strip()
    if not s:
        return None

    if s.casefold() in {"neuvedeny", "neuvedený"}:
        return None

    if s == "0000-00-00":
        return None

    m = _RX_ISO.match(s)
    if m:
        try:
            y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
            if y == 0 or mo == 0 or d == 0:
                return None
            return date(y, mo, d)
        except Exception:
            return None

    m = _RX_DMY.match(s)
    if m:
        try:
            d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
            return date(y, mo, d)
        except Exception:
            return None

    return None


def is_expired(v: Any, *, today: date) -> bool:
    d = parse_date_any(v)
    return (d is not None) and (d < today)


def digits_only(s: str) -> str:
    return re.sub(r"\D+", "", s or "")


def strip_ico(s: Any) -> str:
    d = digits_only(str(s or ""))
    d2 = d.lstrip("0")
    return d2 if d2 else d


def normalize_ascii_str(s: str) -> str:
    return (
        unicodedata.normalize("NFKD", s.casefold())
        .encode("ascii", "ignore")
        .decode("ascii")
    )


def tokens_in_order(haystack_norm: str, tokens: list[str]) -> bool:
    """
    True if each token occurs in haystack in the given order (not necessarily adjacent).
    """
    pos = 0
    for t in tokens:
        if not t:
            continue
        j = haystack_norm.find(t, pos)
        if j < 0:
            return False
        pos = j + len(t)
    return True


def load_ico_set(path: Path) -> set[str]:
    out: set[str] = set()
    if not path.exists():
        return out
    for line in path.read_text(encoding="utf-8").splitlines():
        s = line.strip()
        if not s or s.startswith("#"):
            continue
        ico = strip_ico(s)
        if ico:
            out.add(ico)
    return out


def load_name_patterns(path: Path) -> list[list[str]]:
    """
    Each non-empty line -> list of ascii-normalized tokens.
    Example line: "urad geodez kartograf"
    """
    pats: list[list[str]] = []
    if not path.exists():
        return pats
    for line in path.read_text(encoding="utf-8").splitlines():
        s = line.strip()
        if not s or s.startswith("#"):
            continue
        s_norm = normalize_ascii_str(s)
        toks = [t for t in s_norm.split() if t]
        if toks:
            pats.append(toks)
    return pats


# ----------------------------
# Excel writing (same look as 06)
# ----------------------------

HEADERS = [
    "Dodávateľ",
    "IČO dodávateľa",
    "Názov zmluvy",
    "Číslo zmluvy",
    "Dátum zverejnenia",
    "Dátum uzavretia",
    "Dátum účinnosti",
    "Dátum platnosti do",
    "Objednávateľ",
    "IČO objednávateľa",
    "ID zmluvy",
    "Sken prílohy",
    "Text prílohy",
]

FILL_A = PatternFill("solid", fgColor="E8F5E9")      # light green
FILL_B = PatternFill("solid", fgColor="FFFDE7")      # light yellow
HEADER_FILL = PatternFill("solid", fgColor="ECEFF1") # light gray-blue

LINK_FONT = Font(color="0563C1", underline="single")


def apply_link(cell, url: str) -> None:
    cell.hyperlink = url
    cell.font = LINK_FONT


def apply_row_fill(ws, row: int, max_col: int, fill: PatternFill) -> None:
    for c in range(1, max_col + 1):
        ws.cell(row=row, column=c).fill = fill


def set_comment(ws, row: int, col: int, text: str) -> None:
    cell = ws.cell(row=row, column=col)
    cell.comment = Comment(text, "crz_scraper")


def build_xlsx_from_row_objs(rows: list["RowObj"], out_xlsx: Path) -> None:
    rows = sorted(rows, key=lambda r: (r.contract_ord, r.sub))

    wb = Workbook()
    ws = wb.active
    ws.title = "contracts"

    for col, h in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"

    wrap_cols = {3, 12, 13}
    for col in range(1, len(HEADERS) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 22
        if col in (3, 12, 13):
            ws.column_dimensions[get_column_letter(col)].width = 45
        if col == 11:
            ws.column_dimensions[get_column_letter(col)].width = 12

    row_idx = 2
    group_fill = FILL_A
    cur_ord: int | None = None

    for r in rows:
        if cur_ord is None:
            cur_ord = r.contract_ord
        elif r.contract_ord != cur_ord:
            group_fill = FILL_B if group_fill == FILL_A else FILL_A
            cur_ord = r.contract_ord

        # values
        for col, v in enumerate(r.values, start=1):
            cell = ws.cell(row=row_idx, column=col, value=v)
            cell.alignment = Alignment(vertical="top", wrap_text=(col in wrap_cols))
        apply_row_fill(ws, row_idx, len(HEADERS), group_fill)

        # links: id col=11, scan col=12, text col=13
        if r.links.get("id"):
            apply_link(ws.cell(row=row_idx, column=11), str(r.links["id"]))
        if r.links.get("scan"):
            apply_link(ws.cell(row=row_idx, column=12), str(r.links["scan"]))
        if r.links.get("text"):
            apply_link(ws.cell(row=row_idx, column=13), str(r.links["text"]))

        # comments: B=2 supplier ico, J=10 buyer ico
        if r.comments.get("B"):
            set_comment(ws, row_idx, 2, str(r.comments["B"]))
        if r.comments.get("J"):
            set_comment(ws, row_idx, 10, str(r.comments["J"]))

        row_idx += 1

    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{max(1, row_idx - 1)}"
    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_xlsx)


# ----------------------------
# rows.jsonl reading + grouping
# ----------------------------

@dataclass
class RowObj:
    row_key: str
    contract_ord: int
    sub: int
    parent_cid: int
    values: list[Any]
    links: dict[str, Any]
    comments: dict[str, str]


def read_rows_jsonl(rows_jsonl: Path) -> list[RowObj]:
    out: list[RowObj] = []
    if not rows_jsonl.exists():
        return out

    for line in rows_jsonl.read_text(encoding="utf-8").splitlines():
        if not line.strip():
            continue
        try:
            o = json.loads(line)
        except Exception:
            continue
        if o.get("kind") != "row":
            continue

        try:
            row_key = str(o.get("row_key") or "")
            contract_ord = int(o.get("contract_ord", 10**12))
            sub = int(o.get("sub", 10**12))
            parent_cid = int(o.get("parent_cid", -1))
        except Exception:
            continue

        values = o.get("values") if isinstance(o.get("values"), list) else []
        links = o.get("links") if isinstance(o.get("links"), dict) else {}
        comments = o.get("comments") if isinstance(o.get("comments"), dict) else {}

        out.append(RowObj(
            row_key=row_key,
            contract_ord=contract_ord,
            sub=sub,
            parent_cid=parent_cid,
            values=values,
            links=links,
            comments=comments,
        ))
    return out


def supplier_matches(
    supplier_ico: str,
    supplier_name: str,
    *,
    ico_allow: set[str],
    name_patterns: list[list[str]],
) -> bool:
    """
    Rule:
      - if ICO exists -> must be in ico_allow, otherwise NO match (even if name would match)
      - if ICO missing -> try name pattern match (tokens in order) against dodavatel.txt
    """
    if supplier_ico:
        return supplier_ico in ico_allow

    name_norm = normalize_ascii_str(supplier_name)
    for toks in name_patterns:
        if tokens_in_order(name_norm, toks):
            return True
    return False


# ----------------------------
# main
# ----------------------------

def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--run-dir", type=str, default="", help="filtered_contracts/<sha>/ directory (expects rows.jsonl)")
    ap.add_argument("--rows-jsonl", type=str, default="", help="Explicit rows.jsonl path (overrides --run-dir)")
    ap.add_argument("--out-dir", type=str, default="", help="Output dir (default: <run-dir>/split_tables)")

    ap.add_argument("--ico-file", type=str, default="", help="ICO.txt (default: ICO.txt)")
    ap.add_argument("--dodavatel-file", type=str, default="", help="dodavatel.txt (default: dodavatel.txt)")

    ap.add_argument("--verbose", action="store_true")
    args = ap.parse_args()

    TODAY = today_bratislava()

    project_root = find_project_root()

    if args.rows_jsonl:
        rows_jsonl = Path(args.rows_jsonl)
        run_root = rows_jsonl.parent
    else:
        if not args.run_dir:
            raise SystemExit("Provide --run-dir ... or --rows-jsonl ...")
        run_root = Path(args.run_dir)
        rows_jsonl = run_root / "rows.jsonl"

    if not rows_jsonl.exists():
        raise SystemExit(f"Missing {rows_jsonl}")

    out_dir = Path(args.out_dir) if args.out_dir else (run_root / "split_tables")
    out_dir.mkdir(parents=True, exist_ok=True)

    ico_file = Path(args.ico_file) if args.ico_file else (project_root / "ICO.txt")
    dod_file = Path(args.dodavatel_file) if args.dodavatel_file else (project_root / "dodavatel.txt")

    ico_allow = load_ico_set(ico_file)
    name_patterns = load_name_patterns(dod_file)

    if args.verbose:
        print(f"[load] rows={rows_jsonl}")
        print(f"[load] ICO allow={ico_file} n={len(ico_allow)}")
        print(f"[load] dodavatel patterns={dod_file} n={len(name_patterns)}")

    rows = read_rows_jsonl(rows_jsonl)
    if not rows:
        raise SystemExit("No rows found in rows.jsonl")

    # group by parent contract
    groups: dict[int, list[RowObj]] = {}
    for r in rows:
        groups.setdefault(r.parent_cid, []).append(r)

    buckets: dict[tuple[str, str], list[RowObj]] = {
        ("expired", "match"): [],
        ("expired", "other"): [],
        ("unexpired", "match"): [],
        ("unexpired", "other"): [],
    }

    n_groups = 0
    for parent_cid, grp in groups.items():
        if not grp:
            continue
        n_groups += 1

        # base row = sub==0 if present, else smallest sub
        base = None
        for r in grp:
            if r.sub == 0:
                base = r
                break
        if base is None:
            base = min(grp, key=lambda x: x.sub)

        # robust indexing into values
        vals = base.values or []
        supplier_name = str(vals[0] if len(vals) > 0 else "")
        supplier_ico = strip_ico(vals[1] if len(vals) > 1 else "")
        platnost_do = vals[7] if len(vals) > 7 else ""

        expired = is_expired(platnost_do, today=TODAY)
        match = supplier_matches(
            supplier_ico,
            supplier_name,
            ico_allow=ico_allow,
            name_patterns=name_patterns,
        )

        key = ("expired" if expired else "unexpired", "match" if match else "other")

        # keep entire contract group together
        buckets[key].extend(grp)

    # write 4 xlsx outputs
    def out_path(state: str, vendor: str) -> Path:
        return out_dir / f"{state}_{vendor}.xlsx"

    for (state, vendor), bucket_rows in buckets.items():
        # Sort before writing (writer also sorts, but this keeps group fill stable)
        bucket_rows = sorted(bucket_rows, key=lambda r: (r.contract_ord, r.sub))
        build_xlsx_from_row_objs(bucket_rows, out_path(state, vendor))

    # summary
    def count_contracts(rows_list: list[RowObj]) -> int:
        return len({r.parent_cid for r in rows_list})

    print(f"[ok] wrote to {out_dir}")
    for k in [("unexpired", "match"), ("unexpired", "other"), ("expired", "match"), ("expired", "other")]:
        rows_k = buckets[k]
        print(f"  {k[0]}_{k[1]}: contracts={count_contracts(rows_k)} rows={len(rows_k)} -> {out_path(k[0], k[1]).name}")


if __name__ == "__main__":
    main()
